#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Compare Analyzer 主调度脚本

功能：
1. 读取 compare 输出的 xlsx 文件
2. 调用 11 个 Sheet 解析器提取关键数据
3. 汇总为中间件 JSON
4. 调用 HTML 生成器生成可视化报告
5. 调用 CSV 翻译器生成中文 CSV

用法：
  python main_analyzer.py <xlsx_file_path> [-o <output_dir>]
"""
import os
import sys
import json
import argparse
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook

from parsers.overall_metrics import parse as parse_overall_metrics
from parsers.operator import parse_statistic as parse_operator_statistic
from parsers.operator import parse_compare as parse_operator_compare
from parsers.module import parse_statistic as parse_module_statistic
from parsers.module import parse_compare as parse_module_compare
from parsers.communication import parse as parse_communication
from parsers.memory import parse_statistic as parse_memory_statistic
from parsers.memory import parse_compare as parse_memory_compare
from parsers.kernel import parse_compare as parse_kernel_compare
from parsers.kernel import parse_type as parse_kernel_type
from parsers.api_compare import parse as parse_api_compare


def analyze(xlsx_path: str, output_dir: str = None) -> dict:
    """主分析流程：解析 xlsx → 生成中间件 JSON"""
    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(xlsx_path))
    os.makedirs(output_dir, exist_ok=True)

    print(f"[1/4] 加载 xlsx 文件: {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    print(f"  发现 {len(wb.sheetnames)} 个 Sheet: {wb.sheetnames}")

    print(f"[2/4] 逐 Sheet 解析并提取关键数据...")
    results = {}

    # 1. OverallMetrics
    print("  - OverallMetrics 总体性能比对...")
    results['overall_metrics'] = parse_overall_metrics(wb)

    # 2. OperatorCompareStatistic
    print("  - OperatorCompareStatistic 算子统计...")
    op_stat = parse_operator_statistic(wb)
    results['operator_statistic'] = op_stat

    # 3. OperatorCompare (关联统计页 Top 算子)
    print("  - OperatorCompare 算子明细...")
    top_op_names = [o['name'] for o in op_stat.get('top10', [])] if op_stat.get('available') else None
    results['operator_compare'] = parse_operator_compare(wb, top_op_names)

    # 4. ModuleCompareStatistic
    print("  - ModuleCompareStatistic 模块统计...")
    mod_stat = parse_module_statistic(wb)
    results['module_statistic'] = mod_stat

    # 5. ModuleCompare (关联统计页 Top 模块)
    print("  - ModuleCompare 模块明细...")
    top_mod_names = [m['module_name'] for m in mod_stat.get('top10', [])] if mod_stat.get('available') else None
    results['module_compare'] = parse_module_compare(wb, top_mod_names)

    # 6. CommunicationCompare
    print("  - CommunicationCompare 通信比对...")
    results['communication'] = parse_communication(wb)

    # 7. MemoryCompareStatistic
    print("  - MemoryCompareStatistic 内存统计...")
    mem_stat = parse_memory_statistic(wb)
    results['memory_statistic'] = mem_stat

    # 8. MemoryCompare (关联统计页 Top 算子)
    print("  - MemoryCompare 内存明细...")
    top_mem_names = [o['name'] for o in mem_stat.get('top10', [])] if mem_stat.get('available') else None
    results['memory_compare'] = parse_memory_compare(wb, top_mem_names)

    # 9. KernelCompare
    print("  - KernelCompare Kernel比对...")
    results['kernel_compare'] = parse_kernel_compare(wb)

    # 10. KernelTypeCompare
    print("  - KernelTypeCompare Kernel类型比对...")
    results['kernel_type'] = parse_kernel_type(wb)

    # 11. ApiCompare
    print("  - ApiCompare API比对...")
    results['api_compare'] = parse_api_compare(wb)

    # 生成分析结论
    results['analysis_summary'] = _generate_summary(results)
    results['metadata'] = {
        'source_file': os.path.basename(xlsx_path),
        'analysis_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'total_sheets': len(wb.sheetnames),
        'sheet_names': list(wb.sheetnames),
    }

    # 生成中文翻译 xlsx（传入已加载的 wb 对象，避免重复加载）
    print(f"  生成中文翻译 xlsx...")
    from csv_translator import translate_to_chinese
    translate_dir = os.path.join(output_dir, 'chinese_xlsx')
    chinese_xlsx_path = translate_to_chinese(wb, translate_dir)
    print(f"  中文 xlsx 已保存: {chinese_xlsx_path}")

    wb.close()

    # 保存中间件 JSON
    print(f"[3/4] 保存中间件 JSON...")
    json_path = os.path.join(output_dir, 'compare_analysis_result.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2, default=str)
    print(f"  中间件已保存: {json_path}")

    # 生成 HTML 报告
    print(f"[4/4] 生成 HTML 报告...")
    from html_generator import generate_html_report
    html_path = os.path.join(output_dir, 'compare_analysis_report.html')
    generate_html_report(results, html_path)
    print(f"  HTML 报告已保存: {html_path}")

    print(f"\n[完成] 分析完成！输出目录: {output_dir}")
    print(f"  - 中间件 JSON: {json_path}")
    print(f"  - HTML 报告: {html_path}")
    print(f"  - 中文 xlsx: {chinese_xlsx_path}")

    return results


def _generate_summary(results: dict) -> dict:
    """根据所有 Sheet 的解析结果生成总体分析结论（优化版：分改善亮点和劣化风险）"""
    summary = {
        'improvements': [],
        'risks': [],
        'available_sheets': [],
        'unavailable_sheets': [],
    }

    for key, val in results.items():
        if key in ('analysis_summary', 'metadata'):
            continue
        if isinstance(val, dict):
            if val.get('available'):
                summary['available_sheets'].append(val.get('sheet_name', key))
            else:
                summary['unavailable_sheets'].append(val.get('sheet_name', key))

    # === 总体性能 ===
    om = results.get('overall_metrics', {})
    if om.get('available'):
        e2e = om.get('e2e', {})
        trend = om.get('overall_trend', '持平')
        if e2e:
            summary['improvements' if trend == '改善' else 'risks' if trend == '劣化' else 'improvements'].append(
                f"E2E 总耗时从 {e2e['base']:.2f}ms 变为 {e2e['comp']:.2f}ms"
                f"（{'改善' if e2e['diff'] < 0 else '劣化'} {abs(e2e['diff']):.2f}ms，比率 {e2e['ratio']:.4f}）"
            )

        # 改善亮点
        imp = om.get('primary_improved')
        if imp:
            summary['improvements'].append(
                f"改善最大维度：{imp['name']}（{imp['label']}），"
                f"从 {imp['base_duration']:.2f}ms 降至 {imp['comp_duration']:.2f}ms"
                f"（改善 {abs(imp['diff_duration']):.2f}ms，比率 {imp['diff_ratio']:.4f}）"
            )

        # 劣化风险
        deg = om.get('primary_degraded')
        if deg:
            summary['risks'].append(
                f"劣化最大维度：{deg['name']}（{deg['label']}），"
                f"从 {deg['base_duration']:.2f}ms 升至 {deg['comp_duration']:.2f}ms"
                f"（劣化 {deg['diff_duration']:.2f}ms，比率 {deg['diff_ratio']:.4f}）"
            )

        # Not minimal profiling 警告
        if om.get('has_minimal_warning'):
            summary['risks'].append("E2E 时间存在 Not minimal profiling，可能影响通信和调度耗时判断的准确性")

    # === 通信比对 ===
    comm = results.get('communication', {})
    if comm.get('available'):
        summary['improvements' if comm.get('improved_count', 0) > comm.get('degraded_count', 0) else 'risks'].append(
            f"通信级：共 {comm['total_comm_ops']} 个通信算子，"
            f"改善 {comm['improved_count']} 个，劣化 {comm['degraded_count']} 个"
        )
        # 改善最大的通信算子
        for imp in comm.get('top_improved', [])[:2]:
            desc = (f"通信改善亮点：{imp['comm_name']} 从 {imp['base_total']:.0f}us 降至 {imp['comp_total']:.0f}us"
                    f"（比率 {imp['diff_ratio']:.4f}）")
            if imp.get('is_extreme_improvement'):
                desc += " [极端改善]"
            summary['improvements'].append(desc)

    # === 算子统计 ===
    op_stat = results.get('operator_statistic', {})
    if op_stat.get('available'):
        top1 = op_stat['top10'][0] if op_stat['top10'] else None
        if top1:
            summary['risks'].append(
                f"算子级：共 {op_stat['total_operators']} 个算子，"
                f"劣化 {op_stat['degraded_count']} 个，改善 {op_stat['improved_count']} 个。"
                f"Top1 劣化算子「{top1['name']}」差异 {top1['diff_duration']:.2f}ms。"
                f"Top10 集中度 {op_stat['concentration']:.1f}%。"
            )

    # === 内存统计 ===
    mem = results.get('memory_statistic', {})
    if mem.get('available'):
        if mem.get('all_zero_memory'):
            summary['improvements'].append(
                f"内存级：{mem.get('zero_memory_note', '所有算子内存无差异')}"
            )
        else:
            summary['improvements' if mem['total_diff_memory'] <= 0 else 'risks'].append(
                f"内存级：总内存差异 {mem['total_diff_memory']:.2f}MB，"
                f"增长 {mem['increased_count']} 个算子，减少 {mem['decreased_count']} 个算子"
            )

    # === Kernel 比对 ===
    kc = results.get('kernel_compare', {})
    if kc.get('available'):
        summary['risks'].append(
            f"Kernel级：共 {kc['total_kernels']} 个 Kernel，"
            f"劣化 {kc['degraded_count']} 个（总量 +{kc['total_degraded_us']:.0f}us），"
            f"改善 {kc['improved_count']} 个（总量 -{kc['total_improved_us']:.0f}us），"
            f"净趋势：{kc['net_trend']}"
        )
        for k in kc.get('significant_degraded', [])[:2]:
            summary['risks'].append(
                f"Kernel 劣化风险：{k['kernel']} 总耗时从 {k['base_total']:.0f}us 升至 {k['comp_total']:.0f}us"
                f"（+{k['diff_total']:.0f}us，比率 {k['total_ratio']:.4f}，{k['degradation_type']}）"
            )

    # === API 比对 ===
    api = results.get('api_compare', {})
    if api.get('available'):
        summary['risks'].append(
            f"API级：共 {api['total_apis']} 个 API，"
            f"劣化 {api['degraded_count']} 个（总量 +{api['total_degraded_ms']:.1f}ms），"
            f"改善 {api['improved_count']} 个（总量 -{api['total_improved_ms']:.1f}ms），"
            f"净趋势：{api['net_trend']}。"
            f"其中 {api['pure_duration_change_count']} 个为纯耗时变化，{api['call_count_change_count']} 个涉及调用次数变化"
        )

    # === 模块统计 ===
    mod = results.get('module_statistic', {})
    if mod.get('available') and mod.get('top10'):
        top_mod = mod['top10'][0]
        direction = '改善' if top_mod['diff_total_time'] < 0 else '劣化'
        summary['improvements' if direction == '改善' else 'risks'].append(
            f"模块级：共 {mod['total_modules']} 个模块，"
            f"Top1「{top_mod['module_name']}」{direction} {abs(top_mod['diff_total_time']):.2f}ms"
        )

    return summary


def main():
    parser = argparse.ArgumentParser(description='Compare Analyzer - 解析 compare 输出并生成分析报告')
    parser.add_argument('xlsx_path', type=str, help='compare 输出的 xlsx 文件路径')
    parser.add_argument('-o', '--output', type=str, default=None, help='输出目录（默认与 xlsx 同目录）')
    args = parser.parse_args()

    if not os.path.exists(args.xlsx_path):
        print(f"错误：文件不存在 - {args.xlsx_path}")
        sys.exit(1)

    analyze(args.xlsx_path, args.output)


if __name__ == '__main__':
    main()
