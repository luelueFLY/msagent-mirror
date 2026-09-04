"""公共工具函数：安全类型转换、Sheet 数据读取等"""
from openpyxl.utils import get_column_letter


def safe_float(value, default=0.0):
    try:
        if value is None or str(value).strip() in ('', '/', 'inf', '-inf', 'nan', 'None'):
            if str(value).strip().lower() == 'inf':
                return float('inf')
            return default
        return float(str(value).strip())
    except (ValueError, TypeError):
        return default


def safe_int(value, default=0):
    try:
        if value is None or str(value).strip() in ('', '/', 'None'):
            return default
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return default


def get_sheet_data(wb, sheet_name_keyword, strip=True):
    """通过关键词匹配 Sheet 名，返回 (headers, rows)
    strip=True 时去除首尾空白（默认），strip=False 时保留缩进（用于 OverallMetrics）
    """
    target_ws = None
    for name in wb.sheetnames:
        if sheet_name_keyword.lower() in name.lower():
            target_ws = wb[name]
            break
    if target_ws is None:
        return [], []
    headers = []
    rows = []
    for row_idx, row in enumerate(target_ws.iter_rows(values_only=True), 1):
        if row_idx == 1:
            headers = [str(h).strip() if h is not None else '' for h in row]
            continue
        if strip:
            row_data = [str(v).strip() if v is not None else '' for v in row]
        else:
            row_data = [str(v) if v is not None else '' for v in row]
        if all(v == '' for v in row_data):
            continue
        rows.append(row_data)
    return headers, rows
