#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""中文翻译器

将原版英文 compare 输出 xlsx 翻译为一个完整的中文 xlsx 文件，
每个 Sheet 保留为单独的标签页，表头和 Sheet 名称均翻译为中文。
同时输出独立 CSV 文件。

优化：接收已加载的 Workbook 对象，避免重复加载 xlsx 文件。
"""
import os
import csv as csv_mod
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# 英文 → 中文 表头映射
HEADER_MAP = {
    "Order Id": "序号",
    "Operator Name": "算子名称",
    "Input Shape": "输入Shape",
    "Input Type": "输入类型",
    "Kernel Details": "Kernel详情",
    "Allocated Details": "内存分配详情",
    "Device Duration(us)": "设备耗时(us)",
    "Diff Ratio": "差异比率",
    "Diff Duration(us)": "差异(us)",
    "Diff Size(KB)": "内存差异(KB)",
    "Size(KB)": "内存大小(KB)",
    "Top": "排名",
    "Base Device Duration(ms)": "基准-设备耗时(ms)",
    "Comparison Device Duration(ms)": "比对-设备耗时(ms)",
    "Base Operator Number": "基准-算子数量",
    "Comparison Operator Number": "比对-算子数量",
    "Diff Duration(ms)": "差异(ms)",
    "Base Allocated Duration(ms)": "基准-分配耗时(ms)",
    "Comparison Allocated Duration(ms)": "比对-分配耗时(ms)",
    "Base Allocated Memory(MB)": "基准-分配内存(MB)",
    "Comparison Allocated Memory(MB)": "比对-分配内存(MB)",
    "Diff Memory(MB)": "内存差异(MB)",
    "Communication OP Name": "通信算子名",
    "Task Name": "Task名",
    "Calls": "调用次数",
    "Total Duration(us)": "总耗时(us)",
    "Avg Duration(us)": "平均耗时(us)",
    "Max Duration(us)": "最大耗时(us)",
    "Min Duration(us)": "最小耗时(us)",
    "Module Class": "模块类",
    "Module Name": "模块名",
    "Device Self Time(ms)": "设备自身耗时(ms)",
    "Device Total Time(ms)": "设备总耗时(ms)",
    "Device Self Time Diff(ms)": "自身耗时差异(ms)",
    "Diff Total Ratio": "总耗时比率",
    "Device Total Time Diff(ms)": "总耗时差异(ms)",
    "Device Self Time(us)": "设备自身耗时(us)",
    "Device Total Time(us)": "设备总耗时(us)",
    "Device Self Time Diff(us)": "自身耗时差异(us)",
    "Device Total Time Diff(us)": "总耗时差异(us)",
    "Total Time Ratio": "总耗时比率",
    "Number": "数量",
    "Module Level": "模块层级",
    "Base Call Stack": "基准调用栈",
    "Comparison Call Stack": "比对调用栈",
    "Index": "指标(Index)",
    "Duration(ms)": "耗时(ms)",
    "Duration Ratio": "耗时占比",
    "api name": "API名称",
    "Total Duration(ms)": "总耗时(ms)",
    "Avg Duration(ms)": "平均耗时(ms)",
    "Self Time(ms)": "Self耗时(ms)",
    "Diff Self Ratio": "Self比率",
    "Diff Avg Ratio": "均耗比率",
    "Diff Calls Ratio": "次数比率",
    "Kernel": "Kernel名称",
    "Kernel Type": "Kernel类型",
    "Core Type": "核类型",
}

# Sheet 名称中英文映射
SHEET_NAME_MAP = {
    "OverallMetrics": "总体性能比对",
    "OperatorCompareStatistic": "算子统计比对",
    "OperatorCompare": "算子明细比对",
    "ModuleCompareStatistic": "模块统计比对",
    "ModuleCompare": "模块明细比对",
    "CommunicationCompare": "通信比对",
    "MemoryCompareStatistic": "内存统计比对",
    "MemoryCompare": "内存明细比对",
    "KernelCompare": "Kernel比对",
    "KernelTypeCompare": "Kernel类型比对",
    "ApiCompare": "API比对",
}

# 已知英文字段名列表
KNOWN_HEADERS = [
    'Order Id', 'Operator Name', 'Index', 'Top', 'Communication OP Name',
    'Module Class', 'api name', 'Kernel', 'Kernel Type',
]

# 样式定义
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="7c3aed")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Arial", size=11)
DATA_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="e2e5ea"),
    right=Side(style="thin", color="e2e5ea"),
    top=Side(style="thin", color="e2e5ea"),
    bottom=Side(style="thin", color="e2e5ea"),
)
ZEBRA_FILL = PatternFill(fill_type="solid", fgColor="f6f7f9")


def translate_header(header: str) -> str:
    """翻译单个表头"""
    header = header.strip() if header else ''
    return HEADER_MAP.get(header, header)


def _get_chinese_sheet_name(sheet_name: str) -> str:
    """获取中文 Sheet 名称"""
    for en_key, zh_val in SHEET_NAME_MAP.items():
        if en_key.lower() in sheet_name.lower():
            return zh_val
    return sheet_name


def _is_header_row(row_data: list) -> bool:
    """判断是否为表头行"""
    row_str = ' '.join(str(v) for v in row_data)
    return any(h in row_str for h in KNOWN_HEADERS)


def translate_to_chinese(wb_src, output_dir: str):
    """将已加载的 Workbook 翻译为中文 xlsx + CSV

    优化：直接接收 wb_src 对象，不再重复加载 xlsx 文件。
    写入 xlsx 的同时并行写入 CSV，避免二次遍历。
    """
    os.makedirs(output_dir, exist_ok=True)

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    csv_dir = os.path.join(output_dir, 'chinese_csv')
    os.makedirs(csv_dir, exist_ok=True)

    sheet_count = 0
    total_rows = 0

    for sheet_name in wb_src.sheetnames:
        ws_src = wb_src[sheet_name]
        chinese_name = _get_chinese_sheet_name(sheet_name)

        ws_out = wb_out.create_sheet(title=chinese_name[:31])
        csv_path = os.path.join(csv_dir, f"{chinese_name}.csv")
        csv_file = open(csv_path, 'w', newline='', encoding='utf-8-sig')
        csv_writer = csv_mod.writer(csv_file)

        header_translated = False
        header_row_idx = None
        data_row_count = 0
        max_col_widths = {}

        for row_idx, row in enumerate(ws_src.iter_rows(values_only=True), 1):
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append('')
                else:
                    row_data.append(str(cell))

            if all(v == '' for v in row_data):
                continue

            is_header = not header_translated and _is_header_row(row_data)
            if is_header:
                row_data = [translate_header(h) for h in row_data]
                header_translated = True
                header_row_idx = row_idx

            # 同时写入 xlsx 和 CSV
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_out.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                if is_header:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = HEADER_ALIGN
                else:
                    cell.font = DATA_FONT
                    cell.alignment = DATA_ALIGN
                    if header_row_idx is not None and row_idx > header_row_idx:
                        if data_row_count % 2 == 1:
                            cell.fill = ZEBRA_FILL

                col_letter = get_column_letter(col_idx)
                val_len = len(str(value)) if value else 0
                if col_letter not in max_col_widths or val_len > max_col_widths[col_letter]:
                    max_col_widths[col_letter] = val_len

            csv_writer.writerow(row_data)

            if not is_header:
                data_row_count += 1
            total_rows += 1

        csv_file.close()

        for col_letter, width in max_col_widths.items():
            ws_out.column_dimensions[col_letter].width = min(width + 4, 50)

        freeze_row = (header_row_idx or 1) + 1
        ws_out.freeze_panes = f"A{freeze_row}"

        sheet_count += 1
        print(f"  Sheet: {chinese_name} ({data_row_count} 行)")

    # 保存 xlsx
    output_filename = "compare_chinese_result.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    wb_out.save(output_path)

    print(f"\n  [完成] 中文 xlsx 已生成: {output_path}")
    print(f"  共 {sheet_count} 个 Sheet, {total_rows} 行数据")
    print(f"  CSV 已同步导出到: {csv_dir}/")

    return output_path
